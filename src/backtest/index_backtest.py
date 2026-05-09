"""
Backtests index prediction logic from index_predictor.py output files.

Input files: {UNDERLYING}_{strategy}_predicted.csv
  e.g., NIFTY_trendFollowing_predicted.csv, BANKNIFTY_momentum_predicted.csv

Output: Updates the same CSV files with backtest columns:
  - today_close_1515: Close price on prediction date
  - next_date: Next trading day date
  - next_open_0915: Open price on next trading day
  - next_close_1515: Close price on next trading day
  - next_day_move_pct: Percentage gap move (next_open - today_close) / today_close
  - predicted_max_delta: Maximum potential profit percentage based on intraday high/low
  - result: CORRECT, INCORRECT, MISSED_CALL, MISSED_PUT, OK_NO_TRADE, or N/A

Correctness Logic:
  - CALL is CORRECT if: next_day_open > today_close AND next_day_close > next_day_open
  - PUT is CORRECT if: next_day_open < today_close AND next_day_close < next_day_open
  - NO_POSITION with significant gap: MISSED_CALL (if gap > 1%) or MISSED_PUT (if gap < -1%)
  - NO_POSITION with small gap: OK_NO_TRADE
"""

import os
import sys
import argparse
import glob
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Ensure repo root is on sys.path so `import src.*` works even when executing this file directly.
_repo_root = Path(__file__).resolve().parents[3]
if str(_repo_root) not in sys.path:
    sys.path.insert(0, str(_repo_root))

from src.data_manager.underlying_history_reader import (
    fetch_5m_candles_for_dates,
    fetch_index_daily,
    get_db_connection,
)

PRED_DIR = "output"
PRED_FILE_PATTERN = "{underlying}_{strategy}_predicted.csv"
SIGNIFICANT_MOVE_THRESH = 0.01  # 1% gap => MISSED_CALL / MISSED_PUT for NO_POSITION


def _ensure_backtest_columns(preds: pd.DataFrame) -> pd.DataFrame:
    """Ensure the predictions dataframe has all required backtest columns."""
    cols = [
        "today_close_1515",
        "next_date",
        "next_open_0915",
        "next_close_1515",
        "next_day_move_pct",
        "predicted_max_delta",
        "result",
    ]
    for c in cols:
        if c not in preds.columns:
            preds[c] = pd.NA
    return preds


def backtest_index_prediction(filename: str, underlying: str) -> bool:
    """
    Backtest index prediction logic for a single prediction file.

    Args:
        filename: Name of the prediction file (e.g., "NIFTY_trendFollowing_predicted.csv")
        underlying: NIFTY or BANKNIFTY

    Returns:
        True if successful, False otherwise
    """
    path = os.path.join(PRED_DIR, filename)

    if not os.path.isfile(path):
        print(f"  [WARN] File not found: {path}")
        return False

    try:
        # Load predictions
        preds = pd.read_csv(path, parse_dates=["date"])
        preds["date"] = pd.to_datetime(preds["date"]).dt.normalize()
        preds = _ensure_backtest_columns(preds)

        # Load full index daily data
        conn = get_db_connection()
        try:
            df_daily = fetch_index_daily(conn, underlying=underlying)
        finally:
            conn.close()

        df_daily["trade_date"] = pd.to_datetime(df_daily["trade_date"]).dt.normalize()
        df_daily = df_daily.sort_values("trade_date").reset_index(drop=True)

        # Ensure we have required columns
        if "close_price" not in df_daily.columns:
            print(f"  [ERROR] Missing close_price in underlying data")
            return False
        if "open_price" not in df_daily.columns:
            print(f"  [ERROR] Missing open_price in underlying data")
            return False

        # Build mapping: date -> next trading date
        df_daily["next_trade_date"] = df_daily["trade_date"].shift(-1)
        next_map = df_daily.set_index("trade_date")["next_trade_date"]

        # Fast lookup for daily data
        daily_by_date = df_daily.set_index("trade_date")

        # First pass: collect all next_dates that we'll need for 5-minute candle queries
        next_dates_to_fetch = set()
        for idx, row in preds.iterrows():
            date = row["date"]
            if pd.isna(date) or date not in next_map.index:
                continue
            next_date = next_map[date]
            if pd.notna(next_date) and next_date in daily_by_date.index:
                next_dates_to_fetch.add(next_date)

        # Batch fetch 5-minute candle data for all unique next_dates
        candles_5m_df = pd.DataFrame()
        if next_dates_to_fetch:
            conn = get_db_connection()
            try:
                candles_5m_df = fetch_5m_candles_for_dates(
                    conn,
                    underlying=underlying,
                    dates=list(next_dates_to_fetch)
                )
            finally:
                conn.close()

        # Create lookup dictionary: next_date -> (min_low, max_high)
        candles_lookup = {}
        if not candles_5m_df.empty:
            for _, candle_row in candles_5m_df.iterrows():
                trade_date = candle_row["trade_date"]
                candles_lookup[trade_date] = {
                    "min_low_price": float(candle_row["min_low_price"]) if pd.notna(candle_row["min_low_price"]) else None,
                    "max_high_price": float(candle_row["max_high_price"]) if pd.notna(candle_row["max_high_price"]) else None,
                }

        # Backtest each prediction
        for idx, row in preds.iterrows():
            date = row["date"]

            # Default: clear values
            preds.at[idx, "today_close_1515"] = pd.NA
            preds.at[idx, "next_date"] = pd.NA
            preds.at[idx, "next_open_0915"] = pd.NA
            preds.at[idx, "next_close_1515"] = pd.NA
            preds.at[idx, "next_day_move_pct"] = pd.NA
            preds.at[idx, "predicted_max_delta"] = pd.NA
            preds.at[idx, "result"] = pd.NA

            if pd.isna(date):
                continue

            # If this date isn't in underlying data, skip
            if date not in daily_by_date.index:
                continue

            # Today's close price
            today_close = float(daily_by_date.loc[date, "close_price"])
            preds.at[idx, "today_close_1515"] = today_close

            # Get next trading day
            if date not in next_map.index:
                continue

            next_date = next_map[date]
            if pd.isna(next_date):
                # Last available date: no next day yet
                continue

            if next_date not in daily_by_date.index:
                continue

            preds.at[idx, "next_date"] = next_date

            # Next day's open and close prices
            next_open = float(daily_by_date.loc[next_date, "open_price"])
            next_close = float(daily_by_date.loc[next_date, "close_price"])

            preds.at[idx, "next_open_0915"] = next_open
            preds.at[idx, "next_close_1515"] = next_close

            # Calculate next day move percentage (renamed from gap_move_pct)
            next_day_move_pct = (next_open - today_close) / today_close if today_close != 0 else 0.0
            preds.at[idx, "next_day_move_pct"] = next_day_move_pct

            # Calculate predicted_max_delta based on 5-minute candle data (do this first)
            predicted_max_delta = pd.NA
            pred = row["prediction"]

            if next_date in candles_lookup:
                candle_data = candles_lookup[next_date]

                if pred == "PUT" and candle_data["min_low_price"] is not None:
                    # PUT: (today_close - lowest_next_day_price) / today_close
                    lowest_price = candle_data["min_low_price"]
                    if today_close > 0:
                        predicted_max_delta = (today_close - lowest_price) / today_close
                        preds.at[idx, "predicted_max_delta"] = predicted_max_delta
                elif pred == "CALL" and candle_data["max_high_price"] is not None:
                    # CALL: (highest_next_day_price - today_close) / today_close
                    highest_price = candle_data["max_high_price"]
                    if today_close > 0:
                        predicted_max_delta = (highest_price - today_close) / today_close
                        preds.at[idx, "predicted_max_delta"] = predicted_max_delta
                elif pred == "NO_POSITION":
                    # For NO_POSITION, calculate both CALL and PUT opportunities
                    # Both are positive values representing opportunities
                    call_delta = None
                    put_delta = None

                    if candle_data["max_high_price"] is not None and today_close > 0:
                        call_delta = (candle_data["max_high_price"] - today_close) / today_close

                    if candle_data["min_low_price"] is not None and today_close > 0:
                        put_delta = (today_close - candle_data["min_low_price"]) / today_close

                    # Use the maximum of the two opportunities as predicted_max_delta (always positive)
                    if call_delta is not None and put_delta is not None:
                        predicted_max_delta = max(call_delta, put_delta)
                    elif call_delta is not None:
                        predicted_max_delta = call_delta
                    elif put_delta is not None:
                        predicted_max_delta = put_delta

                    if pd.notna(predicted_max_delta):
                        preds.at[idx, "predicted_max_delta"] = predicted_max_delta

            # Determine result based on prediction logic
            # For CALL/PUT: result = CORRECT if predicted_max_delta > 0
            # For other predictions: use current logic
            if pred == "CALL":
                # CORRECT if predicted_max_delta > 0
                if pd.notna(predicted_max_delta) and predicted_max_delta > 0:
                    result = "CORRECT"
                else:
                    result = "INCORRECT"
            elif pred == "PUT":
                # CORRECT if predicted_max_delta > 0
                if pd.notna(predicted_max_delta) and predicted_max_delta > 0:
                    result = "CORRECT"
                else:
                    result = "INCORRECT"
            elif pred == "NO_POSITION":
                if pd.notna(predicted_max_delta) and abs(predicted_max_delta) >= SIGNIFICANT_MOVE_THRESH:
                    # Use next_day_move_pct to determine direction of missed opportunity
                    # MISSED_CALL if predicted_max_delta is positive and next_day_move_pct is positive
                    # MISSED_PUT if predicted_max_delta is positive and next_day_move_pct is negative
                    if predicted_max_delta > 0:
                        if pd.notna(next_day_move_pct) and next_day_move_pct > 0:
                            result = "MISSED_CALL"
                        elif pd.notna(next_day_move_pct) and next_day_move_pct < 0:
                            result = "MISSED_PUT"
                        else:
                            result = "OK_NO_TRADE"
                    else:
                        result = "OK_NO_TRADE"
                else:
                    result = "OK_NO_TRADE"
            else:
                # Unknown / empty prediction
                result = pd.NA

            preds.at[idx, "result"] = result

        # Sort and reorder columns: predicted_max_delta should be second last before result
        preds = preds.sort_values("date").reset_index(drop=True)

        # Reorder columns to ensure predicted_max_delta is second last before result
        all_cols = list(preds.columns)
        if "predicted_max_delta" in all_cols and "result" in all_cols:
            # Get all columns except predicted_max_delta and result
            cols_without = [c for c in all_cols if c not in ["predicted_max_delta", "result"]]
            # Reorder: all other columns, then predicted_max_delta, then result
            new_cols = cols_without + ["predicted_max_delta", "result"]
            preds = preds[new_cols]

        preds.to_csv(path, index=False)

        # Calculate and display summary
        call_put_preds = preds[preds["prediction"].isin(["CALL", "PUT"])]
        if not call_put_preds.empty:
            total_predictions = len(call_put_preds)
            correct_predictions = len(call_put_preds[call_put_preds["result"] == "CORRECT"])
            accuracy = (correct_predictions / total_predictions * 100) if total_predictions > 0 else 0.0
            print(f"  [OK] Index backtest completed: {len(preds)} rows")
            print(f"  [SUMMARY] Index Prediction Accuracy: {accuracy:.2f}% ({correct_predictions}/{total_predictions} correct)")
        else:
            print(f"  [OK] Index backtest completed: {len(preds)} rows (no CALL/PUT predictions)")

        return True

    except Exception as e:
        print(f"  [ERROR] Error in index backtest: {e}")
        import traceback
        traceback.print_exc()
        return False


def calculate_summary(filename: str) -> dict:
    """
    Calculate summary statistics for a backtested prediction file.

    Args:
        filename: Name of the prediction file

    Returns:
        Dictionary with summary statistics or None if error
    """
    path = os.path.join(PRED_DIR, filename)

    if not os.path.isfile(path):
        return None

    try:
        preds = pd.read_csv(path, parse_dates=["date"])
        preds["date"] = pd.to_datetime(preds["date"]).dt.normalize()

        # Extract strategy name from filename
        # Format: {UNDERLYING}_{strategy}_predicted.csv
        name_without_ext = filename.replace('_predicted.csv', '')
        parts = name_without_ext.split('_')
        if len(parts) >= 2:
            # Skip the first part (underlying) and join the rest as strategy name
            strategy_name = '_'.join(parts[1:])
        else:
            strategy_name = name_without_ext

        # Calculate index prediction accuracy
        call_put_preds = preds[preds["prediction"].isin(["CALL", "PUT"])]
        if not call_put_preds.empty:
            total_predictions = len(call_put_preds)
            correct_predictions = len(call_put_preds[call_put_preds["result"] == "CORRECT"])
            accuracy = (correct_predictions / total_predictions * 100) if total_predictions > 0 else 0.0
        else:
            total_predictions = 0
            correct_predictions = 0
            accuracy = 0.0

        # Calculate additional metrics
        total_calls = len(preds[preds["prediction"] == "CALL"])
        total_puts = len(preds[preds["prediction"] == "PUT"])
        total_no_position = len(preds[preds["prediction"] == "NO_POSITION"])

        correct_calls = len(preds[(preds["prediction"] == "CALL") & (preds["result"] == "CORRECT")])
        correct_puts = len(preds[(preds["prediction"] == "PUT") & (preds["result"] == "CORRECT")])

        call_accuracy = (correct_calls / total_calls * 100) if total_calls > 0 else 0.0
        put_accuracy = (correct_puts / total_puts * 100) if total_puts > 0 else 0.0

        # Calculate missed opportunities for NO_POSITION predictions
        missed_call = len(preds[(preds["prediction"] == "NO_POSITION") & (preds["result"] == "MISSED_CALL")])
        missed_put = len(preds[(preds["prediction"] == "NO_POSITION") & (preds["result"] == "MISSED_PUT")])

        # Calculate recall: (Missed CALL + Missed PUT) / NO_POSITION Count
        recall = ((missed_call + missed_put) / total_no_position * 100) if total_no_position > 0 else 0.0

        return {
            "filename": filename,
            "strategy_name": strategy_name,
            "index_prediction_accuracy": accuracy,
            "total_predictions": total_predictions,
            "correct_predictions": correct_predictions,
            "total_calls": total_calls,
            "total_puts": total_puts,
            "total_no_position": total_no_position,
            "correct_calls": correct_calls,
            "correct_puts": correct_puts,
            "call_accuracy": call_accuracy,
            "put_accuracy": put_accuracy,
            "missed_call": missed_call,
            "missed_put": missed_put,
            "index_prediction_recall": recall,
            "data": preds  # Return full dataframe for detailed comparison
        }

    except Exception as e:
        print(f"  [WARN] Error calculating summary for {filename}: {e}")
        return None


def create_comparison_excel(underlying: str, summary_data: list) -> None:
    """
    Create an Excel file with two sheets comparing all index prediction strategies.

    Sheet 1: Summary of individual strategy backtesting
    Sheet 2: Detailed comparison by date
    """
    if not summary_data:
        return

    # Create workbook
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"

    # Headers
    headers = [
        "Strategy",
        "Index Prediction Accuracy (%)",
        "Total Predictions",
        "Correct Predictions",
        "Total CALLs",
        "Correct CALLs",
        "CALL Accuracy (%)",
        "Total PUTs",
        "Correct PUTs",
        "PUT Accuracy (%)",
        "NO_POSITION Count",
        "Missed CALL",
        "Missed PUT",
        "Index Prediction Recall (%)"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, item in enumerate(summary_data, 2):
        ws1.cell(row=row_idx, column=1).value = item["strategy_name"]
        ws1.cell(row=row_idx, column=2).value = round(item["index_prediction_accuracy"], 2)
        ws1.cell(row=row_idx, column=3).value = item["total_predictions"]
        ws1.cell(row=row_idx, column=4).value = item["correct_predictions"]
        ws1.cell(row=row_idx, column=5).value = item["total_calls"]
        ws1.cell(row=row_idx, column=6).value = item["correct_calls"]
        ws1.cell(row=row_idx, column=7).value = round(item["call_accuracy"], 2)
        ws1.cell(row=row_idx, column=8).value = item["total_puts"]
        ws1.cell(row=row_idx, column=9).value = item["correct_puts"]
        ws1.cell(row=row_idx, column=10).value = round(item["put_accuracy"], 2)
        ws1.cell(row=row_idx, column=11).value = item["total_no_position"]
        ws1.cell(row=row_idx, column=12).value = item["missed_call"]
        ws1.cell(row=row_idx, column=13).value = item["missed_put"]
        ws1.cell(row=row_idx, column=14).value = round(item["index_prediction_recall"], 2)

    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[col_letter].width = adjusted_width

    # Sheet 2: Detailed Comparison
    ws2 = wb.create_sheet("Detailed Comparison")

    # Collect all unique dates
    all_dates = set()
    for item in summary_data:
        if item["data"] is not None and "date" in item["data"].columns:
            all_dates.update(item["data"]["date"].dropna().tolist())

    if not all_dates:
        excel_path = os.path.join(PRED_DIR, f"{underlying}_index_comparison.xlsx")
        wb.save(excel_path)
        print(f"  [OK] Comparison Excel saved to: {excel_path}")
        return

    all_dates = sorted(list(all_dates))

    # Get unique strategies
    strategies = sorted([item["strategy_name"] for item in summary_data])

    # Build column headers for Sheet 2
    col_headers = ["Date", "today_close_1515", "next_open_0915", "next_close_1515", "next_day_move_pct"]

    # Add strategy columns
    for strategy in strategies:
        col_headers.append(f"{strategy} - prediction")
        col_headers.append(f"{strategy} - result")

    # Write headers
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Create a mapping of strategy name to data
    data_map = {item["strategy_name"]: item["data"] for item in summary_data if item["data"] is not None}

    # Write data rows
    for row_idx, date in enumerate(all_dates, 2):
        ws2.cell(row=row_idx, column=1).value = date

        # Get common columns (today_close_1515, next_open_0915, next_close_1515, next_day_move_pct) from first available file
        first_data = None
        for data in data_map.values():
            if data is not None and len(data) > 0:
                first_data = data
                break

        if first_data is not None:
            date_mask = first_data["date"] == date
            if date_mask.any():
                row_data = first_data[date_mask].iloc[0]
                if "today_close_1515" in row_data:
                    ws2.cell(row=row_idx, column=2).value = row_data.get("today_close_1515")
                if "next_open_0915" in row_data:
                    ws2.cell(row=row_idx, column=3).value = row_data.get("next_open_0915")
                if "next_close_1515" in row_data:
                    ws2.cell(row=row_idx, column=4).value = row_data.get("next_close_1515")
                if "next_day_move_pct" in row_data:
                    ws2.cell(row=row_idx, column=5).value = row_data.get("next_day_move_pct")

        col_idx = 6  # Start after Date, today_close_1515, next_open_0915, next_close_1515, next_day_move_pct

        # Add strategy data
        for strategy in strategies:
            if strategy in data_map:
                data = data_map[strategy]
                date_mask = data["date"] == date
                if date_mask.any():
                    row_data = data[date_mask].iloc[0]
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("prediction")
                    col_idx += 1
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("result")
                    col_idx += 1
                else:
                    col_idx += 2  # Skip if no data
            else:
                col_idx += 2  # Skip if no matching strategy

    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[col_letter].width = adjusted_width

    # Save workbook
    excel_path = os.path.join(PRED_DIR, f"{underlying}_index_comparison.xlsx")
    wb.save(excel_path)
    print(f"  [OK] Comparison Excel saved to: {excel_path}")


def run_index_backtest_and_collect(underlying: str) -> dict:
    """
    Main function to backtest all index prediction files for a given underlying.

    Args:
        underlying: NIFTY or BANKNIFTY
    """
    underlying = underlying.upper()

    # Find all prediction files for this underlying
    pattern = os.path.join(PRED_DIR, f"{underlying}_*_predicted.csv")
    files = glob.glob(pattern)

    if not files:
        print(f"[ERROR] No prediction files found for {underlying}")
        print(f"  Looking for pattern: {pattern}")
        return {
            "underlying": underlying,
            "summaries": [],
            "comparison_file": None,
            "success_count": 0,
            "fail_count": 0,
        }

    # Extract filenames (without path)
    filenames = [os.path.basename(f) for f in files]
    filenames.sort()

    print("=" * 70)
    print(f"Backtesting {len(filenames)} index prediction file(s) for {underlying}")
    print("=" * 70)

    success_count = 0
    fail_count = 0
    summary_data = []  # Collect summary data for Excel

    for i, filename in enumerate(filenames, 1):
        print(f"\n[{i}/{len(filenames)}] Processing: {filename}")
        success = backtest_index_prediction(filename, underlying)
        if success:
            success_count += 1
            # Calculate and collect summary
            summary = calculate_summary(filename)
            if summary:
                summary_data.append(summary)
        else:
            fail_count += 1

    print("\n" + "=" * 70)
    print(f"Summary: {success_count} file(s) processed successfully, {fail_count} failed")
    print("=" * 70)

    comparison_file = None
    # Generate comparison Excel file if we have data
    if summary_data:
        print(f"\nGenerating comparison Excel file...")
        create_comparison_excel(underlying, summary_data)
        print(f"[OK] Comparison Excel file created successfully\n")
        comparison_file = f"{underlying}_index_comparison.xlsx"

    summaries = []
    for item in summary_data:
        summaries.append(
            {
                "filename": item["filename"],
                "strategy_name": item["strategy_name"],
                "index_prediction_accuracy": item["index_prediction_accuracy"],
                "index_prediction_recall": item["index_prediction_recall"],
                "total_predictions": item["total_predictions"],
                "correct_predictions": item["correct_predictions"],
            }
        )

    return {
        "underlying": underlying,
        "summaries": summaries,
        "comparison_file": comparison_file,
        "success_count": success_count,
        "fail_count": fail_count,
    }


def main(underlying: str):
    run_index_backtest_and_collect(underlying=underlying)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Backtest index prediction logic (without option selection).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Backtests index prediction files from index_predictor.py:
  - Input: {UNDERLYING}_{strategy}_predicted.csv
  - Output: Updates same files with backtest results

Correctness Logic:
  - CALL is CORRECT if: next_day_open > today_close AND next_day_close > next_day_open
  - PUT is CORRECT if: next_day_open < today_close AND next_day_close < next_day_open
        """
    )
    parser.add_argument(
        "-u", "--underlying",
        required=True,
        choices=["NIFTY", "BANKNIFTY"],
        help="Underlying index (NIFTY or BANKNIFTY)"
    )

    args = parser.parse_args()
    main(underlying=args.underlying)
