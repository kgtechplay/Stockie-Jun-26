# backtest.py - Combined index and option backtest
import os
import sys
import argparse
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from pathlib import Path

# Ensure repo root is on sys.path so `import src.*` works even when executing this file directly.
_repo_root = Path(__file__).resolve().parents[2]
if str(_repo_root) not in sys.path:
    sys.path.insert(0, str(_repo_root))

from src.prediction.options_data_provider import fetch_option_intraday_prices
from src.prediction.underlying_data_provider import (
    fetch_5m_candles_for_dates,
    fetch_index_daily,
    get_db_connection,
)

PRED_DIR = "output"
SIGNIFICANT_MOVE_THRESH = 0.01   # 1% gap => MISSED_CALL / MISSED_PUT for NO_POSITION


def _ensure_backtest_columns(preds: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure the predictions dataframe has all underlying-backtest columns.
    They will be fully overwritten on each run.
    """
    cols = [
        "today_close_1515",
        "next_date",
        "next_open_0915",
        "next_close_1515",
        "next_day_move_pct",
        "predicted_max_delta",
        "result",
    ]
    for c in cols:
        if c not in preds.columns:
            preds[c] = pd.NA
    return preds


def _reorder_columns_after_index_backtest(preds: pd.DataFrame) -> pd.DataFrame:
    """
    Reorder columns so that index backtest columns are inserted between 'prediction' and 'option_trade_date'.
    
    Desired order:
    - date
    - prediction
    - today_close_1515, next_date, next_open_0915, next_close_1515, next_day_move_pct, predicted_max_delta, result (index backtest)
    - option_trade_date, option_instrument_token, ... (option selection columns)
    - option_entry_date, ... (option backtest columns)
    """
    if "prediction" not in preds.columns:
        return preds
    
    # Get all columns
    all_cols = list(preds.columns)
    
    # Define column groups
    index_backtest_cols = ["today_close_1515", "next_date", "next_open_0915", "next_close_1515", "next_day_move_pct", "predicted_max_delta", "result"]
    option_selection_cols = [
        "option_trade_date", "option_instrument_token", "option_tradingsymbol",
        "option_strike", "option_expiry", "option_type", "selection_option_price_1515"
    ]
    option_backtest_cols = [
        "option_entry_date", "option_entry_price_0915", "option_exit_date",
        "option_closing_price_1515", "option_lot_size", "option_pnl_per_contract",
        "option_pnl_per_lot", "option_return_pct", "option_result"
    ]
    
    # Build ordered column list
    ordered_cols = []
    
    # Add columns before 'prediction'
    pred_idx = all_cols.index("prediction") if "prediction" in all_cols else 0
    ordered_cols.extend(all_cols[:pred_idx + 1])  # Include 'prediction'
    
    # Add index backtest columns (only if they exist)
    for col in index_backtest_cols:
        if col in all_cols and col not in ordered_cols:
            ordered_cols.append(col)
    
    # Add option selection columns (only if they exist, in order)
    for col in option_selection_cols:
        if col in all_cols and col not in ordered_cols:
            ordered_cols.append(col)
    
    # Add option backtest columns (only if they exist, in order)
    for col in option_backtest_cols:
        if col in all_cols and col not in ordered_cols:
            ordered_cols.append(col)
    
    # Add any remaining columns that weren't categorized
    remaining_cols = [col for col in all_cols if col not in ordered_cols]
    ordered_cols.extend(remaining_cols)
    
    # Reorder dataframe
    return preds[ordered_cols]


def _reorder_columns_after_option_backtest(preds: pd.DataFrame) -> pd.DataFrame:
    """
    Reorder columns so that option backtest columns are inserted after 'selection_option_price_1515'.
    
    Desired order:
    - date
    - prediction
    - today_close_1515, next_date, next_open_0915, next_close_1515, next_day_move_pct, predicted_max_delta, result (index backtest)
    - option_trade_date, option_instrument_token, option_tradingsymbol, option_strike, 
      option_expiry, option_type, selection_option_price_1515
    - option_entry_date, option_entry_price_0915, ... (option backtest columns)
    """
    if "selection_option_price_1515" not in preds.columns:
        # If selection_option_price_1515 doesn't exist, just use the index backtest reordering
        return _reorder_columns_after_index_backtest(preds)
    
    # Get all columns
    all_cols = list(preds.columns)
    
    # Define column groups
    index_backtest_cols = ["today_close_1515", "next_date", "next_open_0915", "next_close_1515", "next_day_move_pct", "predicted_max_delta", "result"]
    option_selection_cols = [
        "option_trade_date", "option_instrument_token", "option_tradingsymbol",
        "option_strike", "option_expiry", "option_type", "selection_option_price_1515"
    ]
    option_backtest_cols = [
        "option_entry_date", "option_entry_price_0915", "option_exit_date",
        "option_closing_price_1515", "option_lot_size", "option_pnl_per_contract",
        "option_pnl_per_lot", "option_return_pct", "option_result"
    ]
    
    # Build ordered column list
    ordered_cols = []
    
    # Add columns before 'prediction'
    if "prediction" in all_cols:
        pred_idx = all_cols.index("prediction")
        ordered_cols.extend(all_cols[:pred_idx + 1])  # Include 'prediction'
    else:
        # If no prediction column, start from beginning
        ordered_cols.extend([col for col in all_cols if col not in index_backtest_cols + option_selection_cols + option_backtest_cols])
    
    # Add index backtest columns (only if they exist)
    for col in index_backtest_cols:
        if col in all_cols and col not in ordered_cols:
            ordered_cols.append(col)
    
    # Add option selection columns (only if they exist, in order)
    for col in option_selection_cols:
        if col in all_cols and col not in ordered_cols:
            ordered_cols.append(col)
    
    # Add option backtest columns (only if they exist, in order) - these should come after selection_option_price_1515
    for col in option_backtest_cols:
        if col in all_cols and col not in ordered_cols:
            ordered_cols.append(col)
    
    # Add any remaining columns that weren't categorized
    remaining_cols = [col for col in all_cols if col not in ordered_cols]
    ordered_cols.extend(remaining_cols)
    
    # Reorder dataframe
    return preds[ordered_cols]


def _ensure_option_backtest_cols(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "option_entry_date",
        "option_entry_price_0915",
        "option_exit_date",
        "option_closing_price_1515",
        "option_lot_size",
        "option_pnl_per_contract",
        "option_pnl_per_lot",
        "option_return_pct",
        "option_result",
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df


def calculate_and_display_summary(filename: str) -> dict:
    """
    Calculate and display summary metrics for a backtested file.
    
    Metrics:
    1. Index prediction accuracy = % of CORRECT / total predictions
    2. Option selector accuracy = % of PROFIT / total selections when index result was CORRECT
    3. Net profit = sum of option_pnl_per_lot
    """
    path = os.path.join(PRED_DIR, filename)
    
    if not os.path.isfile(path):
        return
    
    try:
        preds = pd.read_csv(path, parse_dates=["date"])
        
        # 1. Index prediction accuracy
        if "result" in preds.columns:
            total_predictions = len(preds[preds["result"].notna()])
            correct_predictions = len(preds[preds["result"] == "CORRECT"])
            
            if total_predictions > 0:
                index_accuracy = (correct_predictions / total_predictions) * 100
            else:
                index_accuracy = 0.0
        else:
            index_accuracy = None
            total_predictions = 0
            correct_predictions = 0
        
        # 2. Option selector accuracy (only for selections where index result was CORRECT)
        if "option_result" in preds.columns and "result" in preds.columns:
            # Filter for rows where index prediction was CORRECT and option was selected
            correct_index_mask = preds["result"] == "CORRECT"
            has_option_result = preds["option_result"].notna()
            correct_with_options = preds[correct_index_mask & has_option_result]
            
            total_correct_selections = len(correct_with_options)
            profitable_selections = len(correct_with_options[correct_with_options["option_result"] == "PROFIT"])
            
            if total_correct_selections > 0:
                option_accuracy = (profitable_selections / total_correct_selections) * 100
            else:
                option_accuracy = 0.0
        else:
            option_accuracy = None
            total_correct_selections = 0
            profitable_selections = 0
        
        # 3. Net profit (sum of option_pnl_per_lot)
        if "option_pnl_per_lot" in preds.columns:
            net_profit = preds["option_pnl_per_lot"].sum()
            if pd.isna(net_profit):
                net_profit = 0.0
            else:
                net_profit = float(net_profit)
        else:
            net_profit = None
        
        # Display summary
        print(f"  [SUMMARY] Summary for {filename}:")
        if index_accuracy is not None:
            print(f"     Index Prediction Accuracy: {index_accuracy:.2f}% ({correct_predictions}/{total_predictions} correct)")
        else:
            print(f"     Index Prediction Accuracy: N/A (no backtest results)")
        
        if option_accuracy is not None:
            print(f"     Option Selector Accuracy: {option_accuracy:.2f}% ({profitable_selections}/{total_correct_selections} profitable when index was correct)")
        else:
            print(f"     Option Selector Accuracy: N/A (no option backtest results)")
        
        if net_profit is not None:
            print(f"     Net Profit: INR {net_profit:,.2f}")
        else:
            print(f"     Net Profit: N/A (no option P&L data)")
        print()
        
        # Return summary data
        return {
            "filename": filename,
            "strategy_combination": filename.replace('.csv', ''),
            "index_prediction_accuracy": index_accuracy if index_accuracy is not None else 0.0,
            "option_selector_accuracy": option_accuracy if option_accuracy is not None else 0.0,
            "net_profit": net_profit if net_profit is not None else 0.0,
            "data": preds  # Return the full dataframe for detailed comparison
        }
        
    except Exception as e:
        print(f"  [WARN] Error calculating summary: {e}")
        return None


def find_prediction_files(underlying: str) -> list:
    """
    Find all CSV files in PRED_DIR that match the pattern:
    {UNDERLYING}_{predictionStrategy}_{selectionStrategy}.csv
    
    This excludes files like {UNDERLYING}_{strategy}_predicted.csv which are
    created by index_predictor.py before option selection.
    
    Args:
        underlying: NIFTY or BANKNIFTY
        
    Returns:
        List of filenames (not full paths) matching the pattern
    """
    underlying = underlying.upper()
    pattern = os.path.join(PRED_DIR, f"{underlying}_*.csv")
    files = glob.glob(pattern)
    
    # Filter to only include files with both prediction and selection strategies
    # Pattern: {UNDERLYING}_{predictionStrategy}_{selectionStrategy}.csv
    # This means the filename should have at least 2 underscores (excluding the one after underlying)
    filenames = []
    for f in files:
        filename = os.path.basename(f)
        # Remove .csv extension
        name_without_ext = filename.replace('.csv', '')
        # Split by underscore
        parts = name_without_ext.split('_')
        
        # Should have: [UNDERLYING, predictionStrategy, selectionStrategy, ...]
        # So at least 3 parts (underlying + 2 strategies)
        if len(parts) >= 3:
            # Check that it's not the old format ending with "_predicted"
            if not name_without_ext.endswith('_predicted'):
                filenames.append(filename)
    
    return sorted(filenames)


def run_index_backtest(filename: str, underlying: str) -> bool:
    """
    Run index backtest for a single prediction file.
    
    Args:
        filename: Name of the prediction file (e.g., "NIFTY_trendFollowing_predicted.csv")
        underlying: NIFTY or BANKNIFTY
        
    Returns:
        True if successful, False otherwise
    """
    path = os.path.join(PRED_DIR, filename)
    
    if not os.path.isfile(path):
        print(f"  [WARN] File not found: {path}")
        return False
    
    try:
        # Load predictions
        preds = pd.read_csv(path, parse_dates=["date"])
        preds["date"] = pd.to_datetime(preds["date"]).dt.normalize()
        preds = _ensure_backtest_columns(preds)

        # Load full index daily data
        conn = get_db_connection()
        try:
            df_daily = fetch_index_daily(conn, underlying=underlying)
        finally:
            conn.close()

        df_daily["trade_date"] = pd.to_datetime(df_daily["trade_date"]).dt.normalize()
        df_daily = df_daily.sort_values("trade_date").reset_index(drop=True)

        # Map column names to what backtest expects
        # fetch_index_daily returns: close_price, open_price -> backtest expects: close_1515, open_0915, open_915
        if "close_price" in df_daily.columns:
            df_daily["close_1515"] = df_daily["close_price"]
        if "open_price" in df_daily.columns:
            df_daily["open_0915"] = df_daily["open_price"]
            df_daily["open_915"] = df_daily["open_price"]  # Also map to open_915 for compatibility

        # Build mapping: date -> next trading date
        df_daily["next_trade_date"] = df_daily["trade_date"].shift(-1)
        next_map = df_daily.set_index("trade_date")["next_trade_date"]

        # Fast lookup for daily data
        daily_by_date = df_daily.set_index("trade_date")

        # First pass: collect all next_dates that we'll need for 5-minute candle queries
        next_dates_to_fetch = set()
        for idx, row in preds.iterrows():
            date = row["date"]
            if pd.isna(date) or date not in next_map.index:
                continue
            next_date = next_map[date]
            if pd.notna(next_date) and next_date in daily_by_date.index:
                next_dates_to_fetch.add(next_date)
        
        # Batch fetch 5-minute candle data for all unique next_dates
        candles_5m_df = pd.DataFrame()
        if next_dates_to_fetch:
            conn = get_db_connection()
            try:
                candles_5m_df = fetch_5m_candles_for_dates(
                    conn,
                    underlying=underlying,
                    dates=list(next_dates_to_fetch)
                )
            finally:
                conn.close()
        
        # Create lookup dictionary: next_date -> (min_low, max_high)
        candles_lookup = {}
        if not candles_5m_df.empty:
            for _, candle_row in candles_5m_df.iterrows():
                trade_date = candle_row["trade_date"]
                candles_lookup[trade_date] = {
                    "min_low_price": float(candle_row["min_low_price"]) if pd.notna(candle_row["min_low_price"]) else None,
                    "max_high_price": float(candle_row["max_high_price"]) if pd.notna(candle_row["max_high_price"]) else None,
                }

        # ---- full recompute for ALL rows ----
        for idx, row in preds.iterrows():
            date = row["date"]

            # Default: clear values; we'll fill if we can compute
            preds.at[idx, "today_close_1515"] = pd.NA
            preds.at[idx, "next_date"] = pd.NA
            preds.at[idx, "next_open_0915"] = pd.NA
            preds.at[idx, "next_close_1515"] = pd.NA
            preds.at[idx, "next_day_move_pct"] = pd.NA
            preds.at[idx, "predicted_max_delta"] = pd.NA
            preds.at[idx, "result"] = pd.NA

            if pd.isna(date):
                continue

            # If this date isn't in underlying data, skip
            if date not in daily_by_date.index:
                continue

            # Today's close
            today_close = float(daily_by_date.loc[date, "close_1515"])
            preds.at[idx, "today_close_1515"] = today_close

            # Next trading day
            if date not in next_map.index:
                continue

            next_date = next_map[date]
            if pd.isna(next_date):
                # last available date: no next day yet
                continue

            if next_date not in daily_by_date.index:
                continue

            preds.at[idx, "next_date"] = next_date

            # Next day's 09:15 open
            next_open = float(daily_by_date.loc[next_date, "open_915"])
            preds.at[idx, "next_open_0915"] = next_open

            # Next day's 15:15 close
            next_close = float(daily_by_date.loc[next_date, "close_1515"])
            preds.at[idx, "next_close_1515"] = next_close

            # Calculate next day move percentage (renamed from gap_move_pct)
            next_day_move_pct = (next_open - today_close) / today_close if today_close != 0 else 0.0
            preds.at[idx, "next_day_move_pct"] = next_day_move_pct

            # Calculate predicted_max_delta based on 5-minute candle data (do this first)
            predicted_max_delta = pd.NA
            pred = row["prediction"]
            
            if next_date in candles_lookup:
                candle_data = candles_lookup[next_date]
                
                if pred == "PUT" and candle_data["min_low_price"] is not None:
                    # PUT: (today_close - lowest_next_day_price) / today_close
                    lowest_price = candle_data["min_low_price"]
                    if today_close > 0:
                        predicted_max_delta = (today_close - lowest_price) / today_close
                        preds.at[idx, "predicted_max_delta"] = predicted_max_delta
                elif pred == "CALL" and candle_data["max_high_price"] is not None:
                    # CALL: (highest_next_day_price - today_close) / today_close
                    highest_price = candle_data["max_high_price"]
                    if today_close > 0:
                        predicted_max_delta = (highest_price - today_close) / today_close
                        preds.at[idx, "predicted_max_delta"] = predicted_max_delta

            # Determine result based on prediction logic
            # For CALL/PUT: result = CORRECT if predicted_max_delta > 0
            # For other predictions: use current logic
            if pred == "CALL":
                # CORRECT if predicted_max_delta > 0
                if pd.notna(predicted_max_delta) and predicted_max_delta > 0:
                    result = "CORRECT"
                else:
                    result = "INCORRECT"
            elif pred == "PUT":
                # CORRECT if predicted_max_delta > 0
                if pd.notna(predicted_max_delta) and predicted_max_delta > 0:
                    result = "CORRECT"
                else:
                    result = "INCORRECT"
            elif pred == "NO_POSITION":
                if abs(next_day_move_pct) >= SIGNIFICANT_MOVE_THRESH:
                    result = "MISSED_CALL" if next_day_move_pct > 0 else "MISSED_PUT"
                else:
                    result = "OK_NO_TRADE"
            else:
                # Unknown / empty prediction
                result = pd.NA

            preds.at[idx, "result"] = result

        preds = preds.sort_values("date").reset_index(drop=True)
        
        # Reorder columns: insert index backtest columns between 'prediction' and 'option_trade_date'
        preds = _reorder_columns_after_index_backtest(preds)
        
        os.makedirs(PRED_DIR, exist_ok=True)
        preds.to_csv(path, index=False)

        print(f"  [OK] Index backtest completed: {len(preds)} rows")
        return True
        
    except Exception as e:
        print(f"  [ERROR] Error in index backtest: {e}")
        import traceback
        traceback.print_exc()
        return False


def run_option_backtest(filename: str, underlying: str, options_view: str | None = None) -> bool:
    """
    Run option backtest for a single prediction file.
    
    Args:
        filename: Name of the prediction file (e.g., "NIFTY_trendFollowing_nearestExpiryATM.csv")
        underlying: NIFTY or BANKNIFTY

        
    Returns:
        True if successful, False otherwise
    """
    path = os.path.join(PRED_DIR, filename)
    
    if not os.path.isfile(path):
        print(f"  [WARN] File not found: {path}")
        return False
    
    try:

        preds = pd.read_csv(path, parse_dates=["date"])
        preds["date"] = preds["date"].dt.normalize()
        preds = _ensure_option_backtest_cols(preds)

        backtest_cols = [
            "option_entry_date",
            "option_entry_price_0915",
            "option_exit_date",
            "option_closing_price_1515",
            "option_lot_size",
            "option_pnl_per_contract",
            "option_pnl_per_lot",
            "option_return_pct",
            "option_result",
        ]
        for c in backtest_cols:
            preds[c] = pd.NA

        mask = (
            preds["prediction"].isin(["CALL", "PUT"])
            & preds["option_instrument_token"].notna()
        )
        needing = preds[mask].copy()
        if needing.empty:
            print(f"  [WARN] No rows with CALL/PUT + option_instrument_token to backtest")
            preds = preds.sort_values("date").reset_index(drop=True)
            preds = _reorder_columns_after_option_backtest(preds)
            os.makedirs(PRED_DIR, exist_ok=True)
            preds.to_csv(path, index=False)
            return True  # Not an error, just no data

        conn = get_db_connection()
        try:
            df_daily = fetch_index_daily(conn, underlying=underlying)
            df_daily["trade_date"] = pd.to_datetime(df_daily["trade_date"]).dt.normalize()
            df_daily = df_daily.sort_values("trade_date").reset_index(drop=True)
            df_daily["next_trade_date"] = df_daily["trade_date"].shift(-1)
            next_map = df_daily.set_index("trade_date")["next_trade_date"]
        finally:
            conn.close()

        entry_info = {}
        entry_dates = []
        tokens = set()

        for idx, row in preds[mask].iterrows():
            pred_date = row["date"]
            if pred_date not in next_map.index:
                continue

            entry_date = next_map[pred_date]
            if pd.isna(entry_date):
                continue

            entry_date = pd.to_datetime(entry_date).normalize()
            entry_info[idx] = entry_date
            entry_dates.append(entry_date)

            try:
                token = int(row["option_instrument_token"])
            except Exception:
                continue

            tokens.add(token)

        if not entry_dates or not tokens:
            print(f"  [WARN] No valid entry dates or tokens to backtest")
            preds = preds.sort_values("date").reset_index(drop=True)
            preds = _reorder_columns_after_option_backtest(preds)
            os.makedirs(PRED_DIR, exist_ok=True)
            preds.to_csv(path, index=False)
            return True  # Not an error, just no data

        start_date = min(entry_dates).date()
        end_date = max(entry_dates).date()

        conn = get_db_connection()
        try:
            prices_df = fetch_option_intraday_prices(
                conn,
                instrument_tokens=tokens,
                start_date=start_date,
                end_date=end_date,
            )
        finally:
            conn.close()

        if prices_df.empty:
            print(f"  [WARN] No option price data found for required tokens/date range")
            preds = preds.sort_values("date").reset_index(drop=True)
            preds = _reorder_columns_after_option_backtest(preds)
            os.makedirs(PRED_DIR, exist_ok=True)
            preds.to_csv(path, index=False)
            return True  # Not an error, just no data

        prices_df["trade_date"] = pd.to_datetime(prices_df["trade_date"]).dt.normalize()

        lookup = {}
        for (token, trade_date), group in prices_df.groupby(["instrument_token", "trade_date"]):
            group_sorted = group.sort_values("snapshot_time")
            entry_row = group_sorted.iloc[0]
            exit_row = group_sorted.iloc[-1]

            entry_price = float(entry_row["option_price"])
            exit_price = float(exit_row["option_price"])
            lot_size = (
                int(group_sorted["lot_size"].iloc[0])
                if "lot_size" in group_sorted.columns and pd.notna(group_sorted["lot_size"].iloc[0])
                else None
            )

            lookup[(int(token), trade_date)] = {
                "entry_price": entry_price,
                "exit_price": exit_price,
                "lot_size": lot_size,
            }

        for idx, row in preds[mask].iterrows():
            if idx not in entry_info:
                continue

            token = int(row["option_instrument_token"])
            entry_date = entry_info[idx]
            key = (token, entry_date)

            if key not in lookup:
                continue

            info = lookup[key]
            entry_price = info["entry_price"]
            exit_price = info["exit_price"]
            lot_size = info["lot_size"]

            preds.at[idx, "option_entry_date"] = entry_date
            preds.at[idx, "option_entry_price_0915"] = entry_price
            preds.at[idx, "option_exit_date"] = entry_date
            preds.at[idx, "option_closing_price_1515"] = exit_price

            if lot_size is not None:
                preds.at[idx, "option_lot_size"] = lot_size

            pnl_per_contract = exit_price - entry_price
            preds.at[idx, "option_pnl_per_contract"] = pnl_per_contract

            if lot_size:
                preds.at[idx, "option_pnl_per_lot"] = pnl_per_contract * lot_size

            if entry_price != 0:
                preds.at[idx, "option_return_pct"] = pnl_per_contract / entry_price

            if pnl_per_contract > 0:
                preds.at[idx, "option_result"] = "PROFIT"
            elif pnl_per_contract < 0:
                preds.at[idx, "option_result"] = "LOSS"
            else:
                preds.at[idx, "option_result"] = "BREAKEVEN"

        preds = preds.sort_values("date").reset_index(drop=True)
        
        # Reorder columns: insert option backtest columns after 'selection_option_price_1515'
        preds = _reorder_columns_after_option_backtest(preds)
        
        os.makedirs(PRED_DIR, exist_ok=True)
        preds.to_csv(path, index=False)

        print(f"  [OK] Option backtest completed: {len(preds[mask])} rows")
        return True
        
    except Exception as e:
        print(f"  [ERROR] Error in option backtest: {e}")
        import traceback
        traceback.print_exc()
        return False


def create_comparison_excel(underlying: str, summary_data: list) -> None:
    """
    Create an Excel file with two sheets comparing all strategy combinations.
    
    Sheet 1: Summary of individual strategy backtesting
    Sheet 2: Detailed comparison by date
    """
    if not summary_data:
        return
    
    # Extract strategy names from filenames
    # Format: {UNDERLYING}_{predictionStrategy}_{selectionStrategy}.csv
    strategy_combinations = {}
    for item in summary_data:
        filename = item["filename"]
        name_parts = filename.replace('.csv', '').split('_')
        if len(name_parts) >= 3:
            pred_strategy = name_parts[1]
            sel_strategy = '_'.join(name_parts[2:])  # Handle multi-word strategy names
            strategy_combinations[filename] = {
                "prediction_strategy": pred_strategy,
                "selection_strategy": sel_strategy
            }
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    
    # Headers
    headers = ["Strategy Combination", "Index Prediction Accuracy (%)", "Option Selector Accuracy (%)", "Net Profit (INR)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, item in enumerate(summary_data, 2):
        ws1.cell(row=row_idx, column=1).value = item["strategy_combination"]
        ws1.cell(row=row_idx, column=2).value = round(item["index_prediction_accuracy"], 2)
        ws1.cell(row=row_idx, column=3).value = round(item["option_selector_accuracy"], 2)
        ws1.cell(row=row_idx, column=4).value = round(item["net_profit"], 2)
    
    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[col_letter].width = adjusted_width
    
    # Sheet 2: Detailed Comparison
    ws2 = wb.create_sheet("Detailed Comparison")
    
    # Collect all unique dates
    all_dates = set()
    for item in summary_data:
        if item["data"] is not None and "date" in item["data"].columns:
            all_dates.update(item["data"]["date"].dropna().tolist())
    
    if not all_dates:
        wb.save(os.path.join(PRED_DIR, f"{underlying}_comparison.xlsx"))
        return
    
    all_dates = sorted(list(all_dates))
    
    # Get unique prediction and selection strategies
    pred_strategies = sorted(set([sc["prediction_strategy"] for sc in strategy_combinations.values()]))
    sel_strategies = sorted(set([sc["selection_strategy"] for sc in strategy_combinations.values()]))
    
    # Build column headers for Sheet 2
    col_headers = ["Date", "today_close_1515", "next_open_0915", "next_close_1515"]
    
    # Add prediction strategy columns
    for pred_strat in pred_strategies:
        col_headers.append(f"{pred_strat} - prediction")
        col_headers.append(f"{pred_strat} - result")
    
    col_headers.append("option_trade_date")
    
    # Add selection strategy columns
    for sel_strat in sel_strategies:
        col_headers.append(f"{sel_strat} - option_tradingsymbol")
        col_headers.append(f"{sel_strat} - option_expiry")
        col_headers.append(f"{sel_strat} - option_pnl_per_lot")
        col_headers.append(f"{sel_strat} - option_result")
    
    # Write headers
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Create a mapping of filename to data
    data_map = {item["filename"]: item["data"] for item in summary_data if item["data"] is not None}
    
    # Write data rows
    for row_idx, date in enumerate(all_dates, 2):
        ws2.cell(row=row_idx, column=1).value = date
        
        # Get common columns (today_close_1515, next_open_0915, next_close_1515) from first available file
        first_data = None
        for data in data_map.values():
            if data is not None and len(data) > 0:
                first_data = data
                break
        
        if first_data is not None:
            date_mask = first_data["date"] == date
            if date_mask.any():
                row_data = first_data[date_mask].iloc[0]
                if "today_close_1515" in row_data:
                    ws2.cell(row=row_idx, column=2).value = row_data.get("today_close_1515")
                if "next_open_0915" in row_data:
                    ws2.cell(row=row_idx, column=3).value = row_data.get("next_open_0915")
                if "next_close_1515" in row_data:
                    ws2.cell(row=row_idx, column=4).value = row_data.get("next_close_1515")
        
        col_idx = 5  # Start after Date, today_close_1515, next_open_0915, next_close_1515
        
        # Add prediction strategy data
        for pred_strat in pred_strategies:
            # Find file with this prediction strategy
            matching_file = None
            for filename, sc in strategy_combinations.items():
                if sc["prediction_strategy"] == pred_strat and filename in data_map:
                    matching_file = filename
                    break
            
            if matching_file and matching_file in data_map:
                data = data_map[matching_file]
                date_mask = data["date"] == date
                if date_mask.any():
                    row_data = data[date_mask].iloc[0]
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("prediction")
                    col_idx += 1
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("result")
                    col_idx += 1
                else:
                    col_idx += 2  # Skip if no data
            else:
                col_idx += 2  # Skip if no matching file
        
        # Add option_trade_date (from first available file with option data)
        for filename, data in data_map.items():
            if data is not None:
                date_mask = data["date"] == date
                if date_mask.any():
                    row_data = data[date_mask].iloc[0]
                    if "option_trade_date" in row_data and pd.notna(row_data.get("option_trade_date")):
                        ws2.cell(row=row_idx, column=col_idx).value = row_data.get("option_trade_date")
                        break
        col_idx += 1
        
        # Add selection strategy data
        for sel_strat in sel_strategies:
            # Find file with this selection strategy
            matching_file = None
            for filename, sc in strategy_combinations.items():
                if sc["selection_strategy"] == sel_strat and filename in data_map:
                    matching_file = filename
                    break
            
            if matching_file and matching_file in data_map:
                data = data_map[matching_file]
                date_mask = data["date"] == date
                if date_mask.any():
                    row_data = data[date_mask].iloc[0]
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("option_tradingsymbol")
                    col_idx += 1
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("option_expiry")
                    col_idx += 1
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("option_pnl_per_lot")
                    col_idx += 1
                    ws2.cell(row=row_idx, column=col_idx).value = row_data.get("option_result")
                    col_idx += 1
                else:
                    col_idx += 4  # Skip if no data
            else:
                col_idx += 4  # Skip if no matching file
    
    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[col_letter].width = adjusted_width
    
    # Save workbook
    excel_path = os.path.join(PRED_DIR, f"{underlying}_comparison.xlsx")
    wb.save(excel_path)
    print(f"  [OK] Comparison Excel saved to: {excel_path}")


def run_e2e_backtest_and_collect(
    underlying: str,
    skip_index: bool = False,
    skip_option: bool = False,
) -> dict:
    """
    Main function to backtest all prediction files for a given underlying.
    
    Args:
        underlying: NIFTY or BANKNIFTY
        skip_index: If True, skip index backtest
        skip_option: If True, skip option backtest
    """
    underlying = underlying.upper()
    
    # Find all matching CSV files
    files = find_prediction_files(underlying)
    
    if not files:
        print(f"[ERROR] No prediction files found for {underlying}")
        print(f"   Searched in: {os.path.join(PRED_DIR, f'{underlying}_*.csv')}")
        return {
            "underlying": underlying,
            "summaries": [],
            "comparison_file": None,
            "success_count": 0,
            "fail_count": 0,
        }
    
    print(f"\n{'='*70}")
    print(f"Backtesting {len(files)} file(s) for {underlying}")
    print(f"{'='*70}\n")
    
    success_count = 0
    fail_count = 0
    summary_data = []  # Collect summary data for Excel
    
    for i, filename in enumerate(files, 1):
        print(f"[{i}/{len(files)}] Processing: {filename}")
        
        # Run index backtest
        if not skip_index:
            index_success = run_index_backtest(filename, underlying)
            if not index_success:
                fail_count += 1
                continue
        else:
            print(f"  ⏭️  Skipping index backtest")
        
        # Run option backtest
        if not skip_option:
            option_success = run_option_backtest(filename, underlying)
            if not option_success:
                fail_count += 1
                continue
        else:
            print(f"  ⏭️  Skipping option backtest")
        
        if (skip_index or index_success) and (skip_option or option_success):
            success_count += 1
            # Calculate and display summary for this file, collect data
            summary = calculate_and_display_summary(filename)
            if summary:
                summary_data.append(summary)
        
        print()  # Empty line between files
    
    print(f"{'='*70}")
    print(f"Summary: {success_count} file(s) processed successfully, {fail_count} failed")
    print(f"{'='*70}\n")
    
    comparison_file = None
    # Generate comparison Excel file if we have data
    if summary_data:
        print(f"Generating comparison Excel file...")
        create_comparison_excel(underlying, summary_data)
        print(f"[OK] Comparison Excel file created successfully\n")
        comparison_file = f"{underlying}_comparison.xlsx"

    summaries = []
    for item in summary_data:
        summaries.append(
            {
                "filename": item["filename"],
                "strategy_combination": item["strategy_combination"],
                "index_prediction_accuracy": item["index_prediction_accuracy"],
                "option_selector_accuracy": item["option_selector_accuracy"],
                "net_profit": item["net_profit"],
            }
        )

    return {
        "underlying": underlying,
        "summaries": summaries,
        "comparison_file": comparison_file,
        "success_count": success_count,
        "fail_count": fail_count,
    }


def main(underlying: str, skip_index: bool = False, skip_option: bool = False):
    run_e2e_backtest_and_collect(
        underlying=underlying,
        skip_index=skip_index,
        skip_option=skip_option,
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Backtest all prediction files for NIFTY or BANKNIFTY.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
This script automatically finds all CSV files matching the underlying pattern and runs
both index and option backtests for each file.

Examples:
  # Backtest all NIFTY files
  python backtest.py -u NIFTY
  
  # Backtest all BANKNIFTY files
  python backtest.py -u BANKNIFTY
  
  # Skip index backtest (only run option backtest)
  python backtest.py -u NIFTY --skip-index
  
  # Skip option backtest (only run index backtest)
  python backtest.py -u NIFTY --skip-option
        """
    )
    parser.add_argument(
        "-u", "--underlying",
        required=True,
        choices=["NIFTY", "BANKNIFTY"],
        help="Underlying index (NIFTY or BANKNIFTY)"
    )
    parser.add_argument(
        "--skip-index",
        action="store_true",
        help="Skip index backtest (only run option backtest)"
    )
    parser.add_argument(
        "--skip-option",
        action="store_true",
        help="Skip option backtest (only run index backtest)"
    )
    
    args = parser.parse_args()
    
    if args.skip_index and args.skip_option:
        print("[ERROR] Cannot skip both index and option backtest")
        sys.exit(1)
    
    main(
        underlying=args.underlying,
        skip_index=args.skip_index,
        skip_option=args.skip_option
    )


